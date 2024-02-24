"""
.xlsx workbook's engine class.

History (most recent first):
2.0.1  2023/3/11    Created.

"""

# Importing modules

# bulitins
import io

# third party
import openpyxl

# self
from ...File.Excel.Sheet import Sheet
from ..Engine import EngineBase, ReaderStream, WriterStream


# Define EngineXLSX class
class EngineXLSX(EngineBase):
    @classmethod
    def read(cls, file: ReaderStream) -> list[Sheet]:
        # Open workbook
        if isinstance(file, bytes):
            content = file
            file = io.BytesIO()
            file.write(content)
            file.seek(0)

        workbook = openpyxl.load_workbook(
            file, read_only=True, keep_vba=False, data_only=True
        )

        # Reading sheets.
        sheets = []
        for s in workbook.worksheets:
            target = Sheet(s.title)
            for num, row in enumerate(s.values):
                target.set_row(num, list(row))
            sheets.append(target)

        return sheets.copy()

    @classmethod
    def save(cls, file: WriterStream, sheets: list[Sheet]):
        # Write to workbook
        workbook = openpyxl.Workbook(write_only=True)  # use write only mode
        for s in sheets:
            # Create a worksheet
            worksheet = workbook.create_sheet(s.name)

            # Writing to worksheet
            for row in s.data:
                worksheet.append(row)
        if file is not None:
            workbook.save(file)
        else:
            stream = io.BytesIO()
            workbook.save(stream)
            return stream.getvalue()
