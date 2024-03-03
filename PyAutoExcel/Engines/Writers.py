import io
from typing import Union

import openpyxl
import xlsxwriter
import xlwt
from xlsxcessive import workbook, xlsx
from xlsxlite.writer import XLSXBook

from .WriterBase import BaseWriter


class XlsxLiteWriter(BaseWriter):
    __engine__ = "xlsxlite"
    _workbook: XLSXBook
    def _setup(self):
        self._workbook = XLSXBook()

    def _write(self):
        for s in self.sheets:
            ws = self._workbook.add_sheet(name=s.name)
            for row in s.data:
                ws.append_row(*row)

    def _output(self, file: Union[str, io.IOBase]):
        self._workbook.finalize(to_file=file, remove_dir=True)



class XlsxCessiveWriter(BaseWriter):
    __engine__ = "xlsxcessive"
    _workbook: workbook.Workbook

    def _setup(self):
        self._workbook = workbook.Workbook()

    def _write(self):
        for s in self.sheets:
            ws = self._workbook.new_sheet(name=s.name)
            for i, row in enumerate(s.data):
                for j, value in enumerate(row):
                    ws.cell(coords=(i, j), value=value)

    def _output(self, file: Union[str, io.IOBase]):
        if isinstance(file, io.IOBase):
            xlsx.save(workbook=self._workbook, filename="", stream=file)
        xlsx.save(workbook=self._workbook, filename=file)


# class XlwtWriter(WriteBook):
#     __engine__ = "xlwt"
#
#     def save_file(self, file_name: str):
#         wb = xlwt.Workbook(encoding="utf-8")
#         for s in self.sheets:
#             ws: xlwt.Worksheet = wb.add_sheet(sheetname=s.name, cell_overwrite_ok=True)
#             for d in s.records:
#                 row, col, value = d
#                 ws.write(r=row, c=col, label=value)
#         wb.save(filename_or_stream=file_name)
#
#     save_io = save_file

class XlwtWriter(BaseWriter):
    __engine__ = 'xlwt'
    _workbook: xlwt.Workbook

    def _setup(self):
        self._workbook = xlwt.Workbook(encoding="utf-8")

    def _write(self):
        for s in self.sheets:
            ws: xlwt.Worksheet = self._workbook.add_sheet(sheetname=s.name, cell_overwrite_ok=True)
            for i, row in enumerate(s.data):
                for j, col in enumerate(row):
                    ws.write(i, j, col)

    def _output(self, file: Union[str, io.IOBase]):
        self._workbook.save(file)


class OpenpyxlWriter(BaseWriter):
    __engine__ = "openpyxl"
    _workbook: openpyxl.Workbook

    def _setup(self):
        self._workbook = openpyxl.Workbook(write_only=True)

    def _write(self):
        for s in self.sheets:
            ws = self._workbook.create_sheet(s.name)
            for row in s.data:
                ws.append(row)

    def _output(self, file: Union[str, io.IOBase]):
        self._workbook.save(file)


class XlsxWriterWriter(BaseWriter):
    __engine__ = "xlsxwriter"
    _workbook: xlsxwriter.Workbook

    def _setup(self):
        self._workbook = xlsxwriter.Workbook(filename="")
        self._workbook.allow_zip64 = True

    def _write(self):
        for s in self.sheets:
            ws = self._workbook.add_worksheet(name=s.name)
            for i, row in enumerate(s.data):
                ws.write_row(i, 0, row)


    def _output(self, file: Union[str, io.IOBase]):
        self._workbook.filename = file
        self._workbook.close()

