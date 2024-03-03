import io

import sxl
import xlrd
import xlsxio
from openpyxl import load_workbook, Workbook
from xlrd.sheet import Sheet as XlrdSheet

from PyAutoExcel.Documents.File.Excel.Sheet import Sheet
from .ReaderBase import BaseReader


class OpenpyxlReader(BaseReader):
    _workbook: Workbook
    __engine__ = "openpyxl"

    def _setup(self):
        if isinstance(self._file, (str, io.BytesIO)):
            self._workbook = load_workbook(self._file, read_only=True)
        elif isinstance(self._file, bytes):
            stream = io.BytesIO(self._file)
            self._workbook = load_workbook(stream, read_only=True)


    def _parse(self):
        for sheet in self._workbook.worksheets:
            ws = Sheet(sheet.title)
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                ws.set_row(i, row)
            self.sheets.append(ws)
            self.sheet_names.append(ws.name)



class XlrdReader(BaseReader):
    _workbook: xlrd.Book
    __engine__ = "xlrd"

    def _setup(self):
        if isinstance(self._file, str):
            self._workbook = xlrd.open_workbook(self._file)
        elif isinstance(self._file, bytes):
            self._workbook = xlrd.open_workbook(file_contents=self._file)
        else:
            self._workbook = xlrd.open_workbook(file_contents=self._file.read())


    def _parse(self):
        for sheet in self._workbook.sheets():
            sheet: XlrdSheet
            ws = Sheet(sheet.name)
            for i in range(sheet.nrows):
                ws.set_row(i, sheet.row_values(i))
            self.sheets.append(ws)
            self.sheet_names.append(ws.name)


class XlsxioReader(BaseReader):
    _workbook: xlsxio.XlsxioReader
    __engine__ = "python-xlsxio"

    def _setup(self):
        if isinstance(self._file, (str, bytes)):
            self._workbook = xlsxio.XlsxioReader(self._file)
        else:
            self._workbook = xlsxio.XlsxioReader(self._file.read())

    def _parse(self):
        for name in self._workbook.get_sheet_names():
            sheet = self._workbook.get_sheet(name)
            ws = Sheet(name)
            for i, row in enumerate(sheet.iter_rows()):
                ws.set_row(i, row)
            self.sheets.append(ws)
            self.sheet_names.append(name)
        self._workbook.close()

# class SxlReader(ReadBook):
#     __engine__ = "sxl"
#
#     def _parse(self):
#         wb = sxl.Workbook(file_obj=self.file_name)
#         ws_count = len(wb.sheets)
#         for i in range(ws_count):
#             ws = wb.sheets[i]
#             self._datas[ws.name] = list(ws.rows)

class SxlReader(BaseReader):
    _workbook: sxl.Workbook
    __engine__ = "sxl"

    def _setup(self):
        if isinstance(self._file, (str, io.IOBase)):
            self._workbook = sxl.Workbook(self._file)
        else:
            self._workbook = sxl.Workbook(io.BytesIO(self._file))


    def _parse(self):
        for i in range(1, len(self._workbook.sheets) // 2 + 1):
            sheet = self._workbook.sheets[i]
            ws = Sheet(sheet.name)
            for j, row in enumerate(sheet.rows):
                ws.set_row(j, row)
            self.sheets.append(ws)
            self.sheet_names.append(ws.name)
